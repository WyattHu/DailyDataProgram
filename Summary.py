import pandas as pd
import datetime
import time
import os
from openpyxl import load_workbook
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from openpyxl.styles import Alignment
import telebot


def summary():
    folder_name = 'DailyData/' + str(datetime.date.today()) + '/Table'

    df_top5 = pd.read_csv(folder_name+'/CryptoTop5-' + str(datetime.date.today()) + '.csv')
    df_top10 = pd.read_csv(folder_name+'/CryptoTop10-' + str(datetime.date.today()) + '.csv')
    df_yahooworldindices = pd.read_csv(folder_name+'/WorldIndices-' + str(datetime.date.today()) + '.csv')
    df_cexvolume = pd.read_csv(folder_name+'/CEX Volume(24h)-' + str(datetime.date.today()) + '.csv')
    df_coinglassbtcetf = pd.read_csv(folder_name+'/Coinglass_BTC-ETF-' + str(datetime.date.today()) + '.csv')
    df_coinglassbtcetf_flowhistory = pd.read_csv(folder_name+'/Coinglass_BTC-ETF-FlowHistory-' + str(datetime.date.today()) + '.csv')
    df_fearandgreed = pd.read_csv(folder_name+'/FearAndGreed-' + str(datetime.date.today()) + '.csv')

    writer = pd.ExcelWriter(folder_name+'/DailyData-' + str(datetime.date.today()) +'.xlsx')
    df_top5.drop(columns=['Unnamed: 0']).to_excel(writer, startrow=0)
    df_yahooworldindices.drop(columns=['Unnamed: 0']).to_excel(writer, startrow=6)
    df_top10.drop(columns=['Unnamed: 0']).to_excel(writer, startrow=13)
    df_cexvolume.drop(columns=['Unnamed: 0']).to_excel(writer, startrow=45)
    df_coinglassbtcetf.drop(columns=['Unnamed: 0']).to_excel(writer, startrow=54)
    df_coinglassbtcetf_flowhistory.drop(columns=['Unnamed: 0']).to_excel(writer, sheet_name='BTC_ETF_FlowHistory')
    df_fearandgreed.drop(columns=['Unnamed: 0']).to_excel(writer, sheet_name='FearAndGreed')

    writer.close()
    time.sleep(0.5)
    folder_name1 = 'DailyData/' + str(datetime.date.today()) + '/Graph/'

    wb = load_workbook(folder_name+'/DailyData-' + str(datetime.date.today()) +'.xlsx')
    sheet = wb.create_sheet("Trend", 1)
    names = ['持仓加权资金费率-BTC','持仓加权资金费率-ETH','USDT杠杆借贷年利率-平均','10年期美债收益率','USDJPY','标普500指数','纳斯达克指数',
           'VIX','美元指数（DXY）', '以太坊价格','黄金价格','Coinbase(COIN)价格', 'WGMI','比特币交易量占比','以太坊交易量占比','比特币现货ETF净流入流出']
    # 将图片保存到Excel工作表中
    sheet.row_dimensions[1].height = 13.3  # 设置行高
    sheet.row_dimensions[2].height = 13.3  # 设置行高
    sheet.row_dimensions[3].height = 53.3  # 设置行高
    for r in range(4,21):
        sheet.row_dimensions[r].height = 30  # 设置行高
    sheet.column_dimensions['F'].width = 37.13  #设置列宽
    sheet.column_dimensions['C'].width = 44.17  #设置列宽
    sheet.column_dimensions['D'].width = 13.24  #设置列宽

    sheet['F3'] = '一个月趋势图'
    sheet['F3'].alignment = Alignment(horizontal='center', vertical='top')

    sheet.add_image(Image(folder_name1 + names[0] + '.jpg'), 'F4')
    sheet.add_image(Image(folder_name1 + names[1] + '.jpg'), 'F5')
    sheet.add_image(Image(folder_name1 + names[2] + '.jpg'), 'F6')
    sheet.add_image(Image(folder_name1 + names[3] + '.jpg'), 'F7')
    sheet.add_image(Image(folder_name1 + names[4] + '.jpg'), 'F9')
    sheet.add_image(Image(folder_name1 + names[5] + '.jpg'), 'F10')
    sheet.add_image(Image(folder_name1 + names[6] + '.jpg'), 'F11')
    sheet.add_image(Image(folder_name1 + names[7] + '.jpg'), 'F12')
    sheet.add_image(Image(folder_name1 + names[8] + '.jpg'), 'F13')
    sheet.add_image(Image(folder_name1 + names[9] + '.jpg'), 'F14')
    sheet.add_image(Image(folder_name1 + names[10] + '.jpg'), 'F15')
    sheet.add_image(Image(folder_name1 + names[11] + '.jpg'), 'F16')
    sheet.add_image(Image(folder_name1 + names[12] + '.jpg'), 'F17')
    sheet.add_image(Image(folder_name1 + names[13] + '.jpg'), 'F18')
    sheet.add_image(Image(folder_name1 + names[14] + '.jpg'), 'F19')
    sheet.add_image(Image(folder_name1 + names[15] + '.jpg'), 'F20')

    sheet['F4'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F5'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F6'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F7'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F8'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F9'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F10'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F11'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F12'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F13'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F14'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F15'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F16'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F17'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F18'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F19'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['F20'].alignment = Alignment(horizontal='center', vertical='center')

    sheet['C4'] = names[0]
    sheet['C5'] = names[1]
    sheet['C6'] = names[2]
    sheet['C7'] = names[3]
    sheet['C9'] = names[4]
    sheet['C10'] = names[5]
    sheet['C11']= names[6]
    sheet['C12']= names[7]
    sheet['C13']= names[8]
    sheet['C14']= names[9]
    sheet['C15']= names[10]
    sheet['C16']= names[11]
    sheet['C17']= names[12]
    sheet['C18']= names[13]
    sheet['C19']= names[14]
    sheet['C20']= names[15]

    sheet['C8'] = '彭博美国企业债总回报指数（LUACTRUU）'
    sheet.add_image(Image('DailyData/LUACTRUU.jpg'), 'F8')
    # sheet['C16'] = '比特币占比'
    # sheet['C17'] = '以太坊占比'

    sheet['D3'] = '最新数值'
    sheet['D3'].alignment = Alignment(horizontal='center', vertical='top')
    btcfee = str(round(float(pd.read_csv(folder_name1 + 'Rowdata/持仓加权资金费率-BTC.csv').iloc[-1]['value'])*100,4))+'%'
    ethfee = str(round(float(pd.read_csv(folder_name1 + 'Rowdata/持仓加权资金费率-ETH.csv').iloc[-1]['value'])*100,4))+'%'
    marginfee = str(round(float(pd.read_csv(folder_name1 + 'Rowdata/USDT杠杆借贷年利率-平均.csv').iloc[-1]['value'])*100,2))+'%'
    inflow = str(float(pd.read_csv(folder_name1 + 'Rowdata/比特币现货ETF净流入流出.csv').iloc[-1]['value']))
    btc_ratio = str(round(100 * float(pd.read_excel(folder_name1+'Rowdata/BTC_ETH_Ratio.xlsx').iloc[-1]['比特币占比'])))+'%'
    eth_ratio = str(round(100 * float(pd.read_excel(folder_name1+'Rowdata/BTC_ETH_Ratio.xlsx').iloc[-1]['以太坊占比'])))+'%'

    usdjpy = str(round(float(pd.read_csv(folder_name1 + 'Rowdata/USDJPY.csv').iloc[-1]['Close']),3))

    sheet['D4'] = btcfee
    sheet['D5'] = ethfee
    sheet['D6'] = marginfee
    sheet['D9'] = usdjpy
    sheet['D18'] = btc_ratio
    sheet['D19'] = eth_ratio
    sheet['D20'] = inflow

    # 保存Excel工作簿
    wb.save(folder_name+'/DailyData-' + str(datetime.date.today()) +'.xlsx')
    bot = telebot.TeleBot("7002822431:AAF_7e6_So04L1Tbzn5uxeoDlcjscx6WduE", parse_mode=None)
    document = open(folder_name+'/DailyData-' + str(datetime.date.today()) +'.xlsx', 'rb')
    bot.send_document(chat_id=-4265387106, document=document)


if __name__ == '__main__':
    summary()
